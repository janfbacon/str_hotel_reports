#!/usr/bin/env python3
"""
process_reports.py — STR Report Ingestion & Database Engine
============================================================
Scans Input Files/ recursively (including inside .zip archives), extracts
12 performance metrics from each weekly STR Excel report's "Glance" worksheet,
and upserts them into STR_Master.xlsx with deduplication on (Inn Code, Date).

After ingestion it runs portfolio analytics and writes performance_brief.txt.

Usage:
    python process_reports.py                       # defaults to Input Files/
    python process_reports.py --input "path/to/dir"  # custom input directory
"""

from __future__ import annotations

import argparse
import io
import logging
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

# ──────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = SCRIPT_DIR.parent / "Input Files"
OUTPUT_DIR = SCRIPT_DIR  # Final Output/
MASTER_FILE = OUTPUT_DIR / "STR_Master.xlsx"
BRIEF_FILE = OUTPUT_DIR / "performance_brief.txt"

# Cell locations on the "Glance" worksheet (openpyxl notation)
METRIC_CELLS: dict[str, str] = {
    # 7-Day % Change
    "MPI_7d_PctChg": "AA12",
    "ARI_7d_PctChg": "AA16",
    "RGI_7d_PctChg": "AA20",
    # 28-Day % Change
    "MPI_28d_PctChg": "AA29",
    "ARI_28d_PctChg": "AA33",
    "RGI_28d_PctChg": "AA37",
    # 7-Day Index
    "MPI_7d_Index": "Z12",
    "ARI_7d_Index": "Z16",
    "RGI_7d_Index": "Z20",
    # 28-Day Index
    "MPI_28d_Index": "Z29",
    "ARI_28d_Index": "Z33",
    "RGI_28d_Index": "Z37",
}

MASTER_COLUMNS = ["Inn Code", "Date"] + list(METRIC_CELLS.keys())

# ──────────────────────────────────────────────────────────────
# Hotel Name / ID → Inn Code Mappings
# ──────────────────────────────────────────────────────────────

# Long hotel names found in "Weekly STAR_<name>-..." filenames.
# Keys are normalised to lowercase with all non-alphanumeric chars stripped.
HOTEL_NAME_MAP: dict[str, str] = {
    "laquintainnsuitesbywyn"
    "dhamchattanoogadowntownsouth": "LQCHA",
    "holidayinnexpresssuites"
    "natchezsouth": "HEZCN",
    "holidayinnexpresssuites"
    "jacksondowntowncoliseum": "JANGM",
    # Spec-mandated display names (kept for completeness / future use)
    "laquintachattanooga": "LQCHA",
    "holidayinnnatchez": "HEZCN",
}

# Numeric property IDs that appear as "Weekly STAR_<id>-..." filenames.
NUMERIC_ID_MAP: dict[str, str] = {
    "62536": "HEZCN",
    "27821": "JANGM",
}

# ──────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────

LOG_FMT = "%(asctime)s  %(levelname)-7s  %(message)s"
LOG_DATE_FMT = "%H:%M:%S"

logger = logging.getLogger("process_reports")


def _setup_logging() -> None:
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter(LOG_FMT, datefmt=LOG_DATE_FMT))
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)


# ──────────────────────────────────────────────────────────────
# 1. Safe workbook loader (strips chart/drawing XML)
# ──────────────────────────────────────────────────────────────

_CHART_DRAWING_RE = re.compile(r"chart|drawing", re.IGNORECASE)
_REL_STRIP_RE = re.compile(
    r'<Relationship[^>]*Target="[^"]*(?:chart|drawing)[^"]*"[^>]*/>',
    re.IGNORECASE,
)


def safe_load_workbook(
    filepath: str | Path,
    *,
    data: bytes | None = None,
) -> openpyxl.Workbook | None:
    """Load an .xlsx workbook, stripping chart/drawing XML to avoid
    openpyxl crashes on files with non-standard chart groupings.

    Pass *data* (raw bytes) when the file lives inside a zip archive
    and has already been read into memory.

    Returns ``None`` on any unrecoverable error.
    """
    source_label = str(filepath)
    try:
        if data is None:
            with open(filepath, "rb") as fh:
                data = fh.read()

        # Re-pack the xlsx zip in memory, omitting chart/drawing entries.
        in_zip = zipfile.ZipFile(io.BytesIO(data))
        buf = io.BytesIO()
        out_zip = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)

        for entry in in_zip.namelist():
            if _CHART_DRAWING_RE.search(entry):
                continue  # skip charts / drawings entirely

            entry_data = in_zip.read(entry)

            # Strip <Relationship> elements that reference charts/drawings
            # from any .rels file so openpyxl doesn't look for them.
            if entry.endswith(".rels"):
                text = entry_data.decode("utf-8")
                text = _REL_STRIP_RE.sub("", text)
                entry_data = text.encode("utf-8")

            out_zip.writestr(entry, entry_data)

        out_zip.close()
        buf.seek(0)

        return openpyxl.load_workbook(buf, data_only=True)

    except Exception:
        logger.warning("Could not load workbook: %s", source_label, exc_info=True)
        return None


# ──────────────────────────────────────────────────────────────
# 2. File discovery (recursive walk + zip extraction)
# ──────────────────────────────────────────────────────────────

# Patterns that identify *Monthly* reports — these do not contain
# weekly index data at the expected cell locations and must be skipped.
_MONTHLY_PATTERN = re.compile(
    r"(Monthly\s*STAR|STARMonthlyReport)", re.IGNORECASE
)


def _is_weekly_report(filename: str) -> bool:
    """Return True if the filename looks like a weekly (not monthly) report."""
    if not filename.lower().endswith(".xlsx"):
        return False
    if filename.startswith("~$"):
        return False  # temp/lock file
    if _MONTHLY_PATTERN.search(filename):
        return False
    return True


def discover_files(
    input_dir: Path,
) -> list[tuple[Path, str, bytes | None]]:
    """Walk *input_dir* recursively, including inside .zip archives.

    Returns a list of ``(filepath, filename, raw_bytes_or_None)`` tuples.
    *raw_bytes* is populated only for files extracted from zip archives
    (since they don't exist on disk).
    """
    results: list[tuple[Path, str, bytes | None]] = []

    for root, _dirs, files in os.walk(input_dir):
        for fname in files:
            full_path = Path(root) / fname

            # Handle zip archives
            if fname.lower().endswith(".zip"):
                logger.info("Extracting zip: %s", full_path.name)
                try:
                    with zipfile.ZipFile(full_path) as zf:
                        for member in zf.namelist():
                            member_name = Path(member).name
                            if _is_weekly_report(member_name):
                                data = zf.read(member)
                                results.append(
                                    (full_path / member, member_name, data)
                                )
                except zipfile.BadZipFile:
                    logger.warning(
                        "Skipping corrupt zip: %s", full_path.name
                    )
                continue

            # Handle loose xlsx files
            if _is_weekly_report(fname):
                results.append((full_path, fname, None))

    logger.info("Discovered %d weekly report file(s)", len(results))
    return results


# ──────────────────────────────────────────────────────────────
# 3. Filename parser → (Inn Code, Date)
# ──────────────────────────────────────────────────────────────

# Pattern A: CODE-YYYYMMDD-USD-E.xlsx
#   e.g. MSYHV-20250907-USD-E.xlsx, JANTW-20250803-USD-E.xlsx
_PAT_CODE_DATE = re.compile(
    r"^([A-Z]{3,10})-(\d{8})-USD-E\.xlsx$", re.IGNORECASE
)

# Pattern B: Weekly STAR_<name_or_id>-YYYYMMDD-USD-E-live.xlsx
#   e.g. Weekly STAR_LaQuinta...-20250810-USD-E-live.xlsx
#   e.g. Weekly STAR_62536-20250810-USD-E-live.xlsx
_PAT_WEEKLY_STAR = re.compile(
    r"^Weekly\s+STAR_(.+?)-(\d{8})-USD-E(?:-live)?\.xlsx$", re.IGNORECASE
)

# Pattern C: DaySTARWeeklyReportUSD_WeekNMonYYYY_[HS-CODE].xlsx
#   e.g. DaySTARWeeklyReportUSD_Week2Jun2025_[HS-HEZCN].xlsx
_PAT_DAYSTAR = re.compile(
    r"^DaySTARWeeklyReportUSD_Week(\d+)([A-Za-z]{3})(\d{4})_\[HS-([A-Z]+)\]\.xlsx$",
    re.IGNORECASE,
)

# Month abbreviation → month number
_MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _normalise_name(name: str) -> str:
    """Strip all non-alphanumeric chars and lowercase."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _resolve_hotel_name(raw_name: str) -> str | None:
    """Map a hotel name (or numeric ID) to an Inn Code, or return None."""
    # Try numeric ID first
    if raw_name.isdigit():
        code = NUMERIC_ID_MAP.get(raw_name)
        if code:
            logger.info(
                "Resolved numeric ID %s → %s (hard-coded fallback)", raw_name, code
            )
        else:
            logger.warning(
                "Unknown numeric property ID '%s' — skipping file", raw_name
            )
        return code

    # Try hotel name lookup
    normed = _normalise_name(raw_name)
    for key, code in HOTEL_NAME_MAP.items():
        if normed == key or normed.startswith(key) or key.startswith(normed):
            return code

    logger.warning("Unmapped hotel name '%s' — skipping file", raw_name)
    return None


def _parse_yyyymmdd(datestr: str) -> datetime | None:
    """Parse an 8-digit date string, returning None for monthly summaries
    (where DD == 00) or invalid dates."""
    if len(datestr) != 8:
        return None
    year = int(datestr[:4])
    month = int(datestr[4:6])
    day = int(datestr[6:8])

    if day == 0:
        # Monthly summary — treat as first day of month so we can still
        # store it if the Glance data is present (some monthly files do
        # have weekly-equivalent cells filled in).
        day = 1

    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def _daystar_date(week_num: int, month_abbr: str, year: int) -> datetime | None:
    """Approximate the date for a DaySTAR 'WeekN<Mon><YYYY>' filename.

    Convention: Week 1 starts on the 1st of the month; each subsequent
    week adds 7 days.
    """
    month = _MONTH_ABBR.get(month_abbr.lower())
    if month is None:
        return None
    day = 1 + (week_num - 1) * 7
    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def parse_filename(filename: str) -> tuple[str, datetime] | None:
    """Extract ``(inn_code, date)`` from *filename*.

    Returns ``None`` when the filename cannot be parsed or the hotel
    cannot be mapped.
    """
    # Pattern A: CODE-YYYYMMDD-USD-E.xlsx
    m = _PAT_CODE_DATE.match(filename)
    if m:
        code = m.group(1).upper()
        dt = _parse_yyyymmdd(m.group(2))
        if dt:
            return code, dt
        return None

    # Pattern B: Weekly STAR_<name_or_id>-YYYYMMDD-...
    m = _PAT_WEEKLY_STAR.match(filename)
    if m:
        raw_name = m.group(1)
        dt = _parse_yyyymmdd(m.group(2))
        if dt is None:
            return None
        code = _resolve_hotel_name(raw_name)
        if code:
            return code, dt
        return None

    # Pattern C: DaySTARWeeklyReportUSD_WeekN<Mon><YYYY>_[HS-CODE].xlsx
    m = _PAT_DAYSTAR.match(filename)
    if m:
        week_num = int(m.group(1))
        month_abbr = m.group(2)
        year = int(m.group(3))
        code = m.group(4).upper()
        dt = _daystar_date(week_num, month_abbr, year)
        if dt:
            return code, dt
        return None

    logger.warning("Unrecognised filename pattern: %s", filename)
    return None


# ──────────────────────────────────────────────────────────────
# 4. Metric extraction from Glance sheet
# ──────────────────────────────────────────────────────────────


def extract_metrics(wb: openpyxl.Workbook) -> dict[str, float | None] | None:
    """Read the 12 metrics from the *Glance* worksheet.

    Returns a dict of metric_name → value, or ``None`` when the sheet
    is missing or all values are None (empty report).
    """
    if "Glance" not in wb.sheetnames:
        logger.warning("Workbook has no 'Glance' sheet (sheets: %s)", wb.sheetnames)
        return None

    ws = wb["Glance"]
    metrics: dict[str, float | None] = {}
    for name, cell_ref in METRIC_CELLS.items():
        raw = ws[cell_ref].value
        if raw is not None:
            try:
                metrics[name] = float(raw)
            except (TypeError, ValueError):
                logger.warning(
                    "Non-numeric value at %s: %r — storing as None", cell_ref, raw
                )
                metrics[name] = None
        else:
            metrics[name] = None

    # If every single metric is None the file likely has a different layout
    if all(v is None for v in metrics.values()):
        logger.warning("All 12 metrics are None — skipping file")
        return None

    return metrics


# ──────────────────────────────────────────────────────────────
# 5. Upsert into STR_Master.xlsx
# ──────────────────────────────────────────────────────────────


def upsert_master(
    new_records: list[dict],
    master_path: Path = MASTER_FILE,
) -> pd.DataFrame:
    """Merge *new_records* into the master Excel file.

    Existing rows with the same ``(Inn Code, Date)`` are overwritten.
    The resulting DataFrame is sorted and saved.
    """
    new_df = pd.DataFrame(new_records, columns=MASTER_COLUMNS)
    new_df["Date"] = pd.to_datetime(new_df["Date"])

    if master_path.exists():
        existing = pd.read_excel(master_path, engine="openpyxl")
        existing["Date"] = pd.to_datetime(existing["Date"])
        # Drop rows that will be replaced by new data
        merge_keys = new_df[["Inn Code", "Date"]]
        mask = existing.set_index(["Inn Code", "Date"]).index.isin(
            merge_keys.set_index(["Inn Code", "Date"]).index
        )
        kept = existing[~mask]
        combined = pd.concat([kept, new_df], ignore_index=True)
        n_replaced = mask.sum()
        if n_replaced:
            logger.info("Overwrote %d existing record(s) on upsert", n_replaced)
    else:
        combined = new_df

    combined.sort_values(["Inn Code", "Date"], inplace=True)
    combined.reset_index(drop=True, inplace=True)

    # Save
    combined.to_excel(master_path, index=False, engine="openpyxl")
    logger.info(
        "Saved %d record(s) to %s", len(combined), master_path.name
    )

    return combined


# ──────────────────────────────────────────────────────────────
# 6. Analytics & Brief Generation
# ──────────────────────────────────────────────────────────────


def generate_brief(df: pd.DataFrame, output_path: Path = BRIEF_FILE) -> None:
    """Analyse the master dataset and write performance_brief.txt."""
    if df.empty:
        logger.warning("No data to analyse — skipping brief generation")
        return

    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"])
    df.sort_values(["Inn Code", "Date"], inplace=True)

    lines: list[str] = []
    lines.append("=" * 68)
    lines.append("  HERMES HOSPITALITY — STR PORTFOLIO PERFORMANCE BRIEF")
    lines.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("=" * 68)

    # --- Portfolio Momentum ---
    # Count of properties whose latest 7-Day RGI % Change improved vs prior week
    lines.append("\n─── PORTFOLIO MOMENTUM ───")
    improving = 0
    declining = 0
    stable = 0
    hotels = sorted(df["Inn Code"].unique())

    for code in hotels:
        hdf = df[df["Inn Code"] == code].sort_values("Date")
        if len(hdf) < 2:
            continue
        latest = hdf.iloc[-1]["RGI_7d_PctChg"]
        prior = hdf.iloc[-2]["RGI_7d_PctChg"]
        if latest is not None and prior is not None:
            if latest > prior:
                improving += 1
            elif latest < prior:
                declining += 1
            else:
                stable += 1

    total_with_history = improving + declining + stable
    lines.append(
        f"  Properties with improving 7-Day RGI % Change: "
        f"{improving} of {total_with_history}"
    )
    lines.append(f"  Declining: {declining}  |  Stable: {stable}")

    # --- Historical Trends (26-week high/low on 28-Day RGI Index) ---
    lines.append("\n─── HISTORICAL TRENDS (26-Week Window) ───")
    for code in hotels:
        hdf = df[df["Inn Code"] == code].sort_values("Date")
        vals = hdf["RGI_28d_Index"].dropna()
        if len(vals) < 2:
            continue
        latest_val = vals.iloc[-1]
        window = vals.tail(26)
        if latest_val == window.max():
            lines.append(f"  ▲ {code}: 28-Day RGI Index at 26-week HIGH ({latest_val:.2f})")
        elif latest_val == window.min():
            lines.append(f"  ▼ {code}: 28-Day RGI Index at 26-week LOW ({latest_val:.2f})")

    # --- Urgent Alerts (WoW swing > ±5% on 7-Day RGI % Change) ---
    lines.append("\n─── URGENT ALERTS (Week-over-Week Swing > ±5%) ───")
    alert_found = False
    for code in hotels:
        hdf = df[df["Inn Code"] == code].sort_values("Date")
        if len(hdf) < 2:
            continue
        latest = hdf.iloc[-1]["RGI_7d_PctChg"]
        prior = hdf.iloc[-2]["RGI_7d_PctChg"]
        if latest is not None and prior is not None:
            swing = latest - prior
            if abs(swing) > 5:
                direction = "⚠ SURGE" if swing > 0 else "⚠ DROP"
                lines.append(
                    f"  {direction}  {code}: 7-Day RGI % Change swung "
                    f"{swing:+.2f}pp  ({prior:.2f}% → {latest:.2f}%)"
                )
                alert_found = True

    if not alert_found:
        lines.append("  No urgent swings detected this period.")

    # --- Rankings (Top 3 & Bottom 3 by 28-Day RGI % Change) ---
    lines.append("\n─── RANKINGS (by Latest 28-Day RGI % Change) ───")
    latest_rows = df.sort_values("Date").groupby("Inn Code").tail(1)
    ranked = latest_rows.dropna(subset=["RGI_28d_PctChg"]).sort_values(
        "RGI_28d_PctChg", ascending=False
    )

    lines.append("  Top Performers:")
    for i, (_, row) in enumerate(ranked.head(3).iterrows(), 1):
        lines.append(
            f"    {i}. {row['Inn Code']:6s}  {row['RGI_28d_PctChg']:+.2f}%"
        )

    lines.append("  Bottom Performers:")
    for i, (_, row) in enumerate(ranked.tail(3).iterrows(), 1):
        lines.append(
            f"    {i}. {row['Inn Code']:6s}  {row['RGI_28d_PctChg']:+.2f}%"
        )

    lines.append("\n" + "=" * 68)
    lines.append("  End of Brief")
    lines.append("=" * 68)

    brief_text = "\n".join(lines) + "\n"
    output_path.write_text(brief_text, encoding="utf-8")
    logger.info("Wrote performance brief → %s", output_path.name)
    print("\n" + brief_text)


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingest STR weekly reports and build STR_Master.xlsx"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Root directory to scan for STR reports (default: Input Files/)",
    )
    args = parser.parse_args()
    input_dir: Path = args.input.resolve()

    _setup_logging()

    print()
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║   HERMES STR REPORT PROCESSOR                              ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()

    if not input_dir.exists():
        logger.error("Input directory not found: %s", input_dir)
        sys.exit(1)

    logger.info("Scanning: %s", input_dir)

    # Step 1 — Discover files
    file_list = discover_files(input_dir)
    if not file_list:
        logger.error("No report files found in %s", input_dir)
        sys.exit(1)

    # Step 2 — Ingest
    records: list[dict] = []
    skipped = 0
    errors = 0

    for filepath, filename, raw_data in file_list:
        parsed = parse_filename(filename)
        if parsed is None:
            skipped += 1
            continue

        inn_code, report_date = parsed

        wb = safe_load_workbook(filepath, data=raw_data)
        if wb is None:
            errors += 1
            continue

        metrics = extract_metrics(wb)
        if metrics is None:
            skipped += 1
            continue

        record = {"Inn Code": inn_code, "Date": report_date}
        record.update(metrics)
        records.append(record)

        logger.info(
            "  ✓ %-6s  %s  (%s)",
            inn_code,
            report_date.strftime("%Y-%m-%d"),
            filename,
        )

    # Step 3 — Summary
    print()
    logger.info(
        "Ingestion complete: %d ingested, %d skipped, %d errors",
        len(records),
        skipped,
        errors,
    )

    if not records:
        logger.error("No records to write — aborting")
        sys.exit(1)

    # Step 4 — Upsert
    master_df = upsert_master(records)

    # Step 5 — Show summary table
    print()
    summary = master_df.groupby("Inn Code").agg(
        Records=("Date", "count"),
        Earliest=("Date", "min"),
        Latest=("Date", "max"),
    )
    print("┌──────────────────────────────────────────────────┐")
    print("│  MASTER DATABASE SUMMARY                         │")
    print("├──────────┬─────────┬────────────┬────────────────┤")
    print("│ Inn Code │ Records │  Earliest  │    Latest      │")
    print("├──────────┼─────────┼────────────┼────────────────┤")
    for code, row in summary.iterrows():
        print(
            f"│ {code:8s} │ {row['Records']:7d} │ "
            f"{row['Earliest'].strftime('%Y-%m-%d')} │ "
            f"{row['Latest'].strftime('%Y-%m-%d')}     │"
        )
    print("└──────────┴─────────┴────────────┴────────────────┘")
    print(f"\n  Total records in STR_Master.xlsx: {len(master_df)}")

    # Step 6 — Analytics & Brief
    generate_brief(master_df)


if __name__ == "__main__":
    main()
